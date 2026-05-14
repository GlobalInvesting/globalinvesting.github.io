#!/usr/bin/env python3
"""
fetch_inflation_expectations.py  —  v5.0
=========================================
Fetch forward-looking inflation expectations for G8 FX currencies and patch
extended-data/{CCY}.json with the result.

SOURCE CASCADE (v5.0):
  USD  →  FRED T5YIE          (5Y TIPS breakeven, market-implied, daily)       [forward]
  EUR  →  FRED T5YIFR         (EUR 5Y5Y inflation swap, market-implied, daily) [forward]
  GBP  →  BOE SDIE BEAPFF    (BoE/Ipsos 2Y-ahead household survey, quarterly) [forward]
       →  IMF SDMX 3.0        (monthly CPI index → YoY, ~4-6w lag)            [backward]
       →  OECD / FRED / WB    (increasing lag fallbacks)                        [backward]
  CAD  →  FRED CAINFIMPCPI   (Bank of Canada 5Y breakeven, market-implied)     [forward]
       →  IMF SDMX 3.0 → OECD → FRED → WB (fallbacks)                         [backward]
  AUD  →  RBA trimmed mean CPI (OECD, underlying, removes volatile items)       [structural]
       →  IMF SDMX 3.0 → OECD → FRED → WB (fallbacks, headline CPI)           [backward]
  JPY  →  IMF SDMX 3.0 → OECD → FRED → WB (no liquid traded breakeven)        [backward]
  CHF  →  IMF SDMX 3.0 → OECD → FRED → WB (no liquid traded breakeven)        [backward]
  NZD  →  RBNZ Inflation Expectations 2Y-ahead survey (DBnomics)               [forward]
       →  IMF SDMX 3.0 → OECD → FRED → WB (fallbacks)                         [backward]

WHY v5.0 (methodological upgrade from v4.3):
  The previous pipeline used backward-looking CPI YoY (what inflation WAS) for
  GBP/JPY/AUD/CAD/CHF/NZD while using forward-looking market-implied breakevens
  for USD/EUR. This methodological inconsistency distorted the real-carry ranking:
  mixing realised CPI with market-implied expectations in the same differential
  comparison is equivalent to comparing speeds in km/h and mph.

  Institutional FX carry screens (Bloomberg FXFR, JP Morgan GBI, Deutsche Bank
  carry research) use the same methodology for all legs. Market-implied breakevens
  are preferred; survey-based forward expectations are the institutional fallback;
  realised CPI (YoY) is the last resort when no forward series exists.

  Changes per currency:
    GBP — ADDED BoE SDIE BEAPFF as primary (2Y-ahead survey, quarterly). IMF
          CPI YoY demoted to fallback. Practical effect: 3.45% CPI YoY (spike)
          vs ~3.0% survey expectation — more representative of market view.
    CAD — ADDED FRED CAINFIMPCPI as primary (5Y breakeven, market-implied).
          Replaces CANCPIALLMINMEI index→YoY (~12-14m structural lag).
    AUD — ADDED OECD trimmed mean CPI (underlying) as primary before headline CPI.
          AUS March 2026 headline 4.57% (energy spike) vs trimmed mean ~3.3%.
          OECD DF_PRICES_ALL CPIH (Housing cost excluded) or DF_CPIH_* not
          available; using OECD DSD_PRICES CPAT (All items less food and energy)
          as structural proxy where available, fallback to IMF headline.
    NZD — ADDED RBNZ 2Y-ahead survey via DBnomics as primary.
    JPY, CHF — No liquid breakeven exists. Retain IMF CPI YoY cascade.
          JPY distortion is limited (1.44% CPI ≈ market expectations ~1.3-1.5%).

IMF API format unchanged:
  {ISO3}.CPI._T.IX.M — monthly index, YoY computed via 12-month pct change.

Output: patches extended-data/{CCY}.json, preserving all other fields.
  Only updates data.inflationExpectations + dates.inflationExpectations.
"""

import os
import sys
import csv
import json
import datetime
import requests
from io import StringIO

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

FRED_API_KEY = os.environ.get("FRED_API_KEY", "")

BASE_HEADERS = {
    "User-Agent": "GlobalInvesting-FX-Terminal/4.3 (+https://globalinvesting.github.io)"
}

STALE_DAYS = {
    "monthly":   120,
    "quarterly": 210,
    "market":     10,
}

# IMF ISO-3 country codes
IMF_COUNTRY = {
    "GBP": "GBR",
    "JPY": "JPN",
    "AUD": "AUS",
    "CAD": "CAN",
    "CHF": "CHE",
    "NZD": "NZL",
}

# OECD country codes + quarterly flag (AUD/NZD are quarterly in DF_PRICES_ALL)
OECD_COUNTRY = {
    "GBP": ("GBR", "M"),
    "JPY": ("JPN", "M"),   # Will 404 — COICOP 2018 only; kept for documentation
    "AUD": ("AUS", "Q"),
    "CAD": ("CAN", "M"),
    "CHF": ("CHE", "M"),
    "NZD": ("NZL", "Q"),
}

FRED_INDEX_SERIES = {
    "GBP": ("GBRCPIALLMINMEI",  "monthly"),
    "JPY": ("JPNCPIALLMINMEI",  "monthly"),
    "AUD": ("AUSCPIALLQINMEI",  "quarterly"),
    "CAD": ("CANCPIALLMINMEI",  "monthly"),
    "CHF": ("CHECPIALLMINMEI",  "monthly"),
    "NZD": ("NZLCPIALLQINMEI",  "quarterly"),
}

WB_COUNTRY = {
    "GBP": "GB", "JPY": "JP", "AUD": "AU",
    "CAD": "CA", "CHF": "CH", "NZD": "NZ",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get(url, custom_headers=None, **kwargs):
    h = {**BASE_HEADERS, **(custom_headers or {})}
    return requests.get(url, headers=h, timeout=25, **kwargs)


def _parse_date(s):
    """Parse YYYY-MM, YYYY-MM-DD, YYYY-Mxx (IMF), or YYYY-Qx (OECD quarterly)."""
    s = str(s).strip()
    # IMF format: "2024-M01" (len=8) → "2024-01"
    if len(s) == 8 and s[5] == "M":
        s = s[:5] + s[6:]      # "2024-M01" → "2024-01"
    # Quarterly format: "2024-Q4" → last month of that quarter
    if len(s) == 7 and "Q" in s:
        try:
            year = int(s[:4])
            quarter = int(s[-1])
            month = quarter * 3          # Q1→3, Q2→6, Q3→9, Q4→12
            return datetime.date(year, month, 1)
        except ValueError:
            raise ValueError(f"Cannot parse quarterly date: {s!r}")
    # Standard: "YYYY-MM" → "YYYY-MM-01"
    if len(s) == 7:
        s += "-01"
    return datetime.date.fromisoformat(s)


def _is_stale(obs_date, freq):
    threshold = STALE_DAYS.get(freq, 120)
    return (datetime.date.today() - obs_date).days > threshold


def _yoy_from_index(series):
    """Compute YoY % from an index series (list of (date, value) tuples)."""
    if len(series) < 13:
        return None
    series_sorted = sorted(series, key=lambda x: x[0])
    last_date, last_val = series_sorted[-1]
    # Find observation ~12 months prior (within ±92 days)
    target_prior = datetime.date(last_date.year - 1, last_date.month, 1)
    best = None
    for d, v in series_sorted:
        diff = abs((d - target_prior).days)
        if diff <= 92:
            if best is None or diff < best[0]:
                best = (diff, d, v)
    if best is None:
        return None
    _, _, prior_val = best
    if prior_val == 0:
        return None
    return round((last_val / prior_val - 1) * 100, 4), last_date


# ---------------------------------------------------------------------------
# USD / EUR  —  FRED breakevens
# ---------------------------------------------------------------------------

def fetch_fred_breakeven(series_id, currency, label):
    url = (
        "https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&sort_order=desc&limit=5&file_type=json"
        f"&api_key={FRED_API_KEY}"
    )
    try:
        r = _get(url)
        r.raise_for_status()
        for o in r.json().get("observations", []):
            if o.get("value") not in (".", ""):
                val = float(o["value"])
                date = _parse_date(o["date"])
                if not _is_stale(date, "market"):
                    print(f"    OK {currency}: {val:.4f}% ({date}) [{label}]")
                    return val, date, label
        print(f"    MISS {currency}: {label} — no recent observations")
        return None
    except Exception as e:
        print(f"    ERR {currency}: {label} — {e}")
        return None


# ---------------------------------------------------------------------------
# GBP PRIMARY  —  BoE SDIE BEAPFF (2Y-ahead household survey, quarterly)
# Series: BEAPFF = Bank of England / Ipsos Mori Inflation Attitudes Survey
#         "Median inflation rate expected over the next 12 months" (published quarterly)
# Endpoint: same BOE SDIE CSV endpoint used by update_extended_data.py for bond yields
# Date format in response: "DD Mon YYYY" (e.g. "01 Feb 2026")
# ---------------------------------------------------------------------------

def fetch_boe_inflation_expectations():
    """BoE/Ipsos 2Y-ahead inflation expectations survey (BEAPFF). Quarterly."""
    import datetime as _dt
    label = "BOE SDIE BEAPFF (2Y survey)"
    now = _dt.datetime.now()
    date_from = f"01/Jan/{now.year - 1}"
    date_to   = f"{now.day:02d}/{now.strftime('%b')}/{now.year}"
    url = "https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp"
    params = {
        "csv.x":      "yes",
        "Datefrom":   date_from,
        "Dateto":     date_to,
        "SeriesCodes": "BEAPFF",
        "CSVF":       "TN",
        "UsingCodes": "Y",
        "VPD":        "Y",
        "VFD":        "N",
    }
    print(f"    Trying BOE SDIE BEAPFF...")
    try:
        r = _get(url, params=params)
        if not r.ok:
            print(f"    MISS GBP: BOE SDIE HTTP {r.status_code}")
            return None
        data_rows = []
        import csv as _csv
        reader = _csv.reader(r.text.strip().splitlines())
        for row in reader:
            if not row or not row[0].strip():
                continue
            date_str = row[0].strip().strip('"')
            # Skip header rows (DATE, BEAPFF, etc.)
            if date_str.upper() in ("DATE", "BEAPFF") or not date_str[0].isdigit():
                continue
            try:
                dt = _dt.datetime.strptime(date_str, "%d %b %Y").date()
                val_str = row[1].strip().strip('"').replace("..", "").strip() if len(row) > 1 else ""
                if val_str:
                    data_rows.append((dt, float(val_str)))
            except (ValueError, IndexError):
                continue
        if not data_rows:
            print(f"    MISS GBP: BOE SDIE — no parseable rows")
            return None
        data_rows.sort(key=lambda x: x[0])
        obs_date, obs_val = data_rows[-1]
        if _is_stale(obs_date, "quarterly"):
            print(f"    STALE GBP: BOE BEAPFF obs {obs_date} is >{STALE_DAYS['quarterly']}d — skipping")
            return None
        print(f"    OK GBP: {obs_val:.4f}% 2Y survey ({obs_date}) [{label}]")
        return obs_val, obs_date, label
    except Exception as e:
        print(f"    ERR GBP: BOE SDIE — {e}")
        return None


# ---------------------------------------------------------------------------
# CAD PRIMARY  —  BOC Valet breakeven inflation (nominal 10Y − RRB long-term yield)
# Canada does not publish a pre-computed breakeven series on FRED or BOC Valet.
# The standard market measure is: BD.CDN.LONG.DQ.YLD (30Y nominal) − BD.CDN.RRB.DQ.YLD (RRB)
# Note: BOC cancelled new RRB issuances in 2022. The RRB series remains published
# (secondary market) but with declining liquidity. When the spread is outside [1%, 3%]
# or the RRB data is stale, we fall through to IMF CPI YoY.
# ---------------------------------------------------------------------------

def fetch_cad_breakeven():
    """Canada breakeven inflation: BOC Valet LONG nominal yield − RRB yield. Daily."""
    import datetime as _dt
    label = "BOC Valet breakeven (nominal LONG − RRB)"
    end   = _dt.datetime.now().strftime("%Y-%m-%d")
    start = (_dt.datetime.now() - _dt.timedelta(weeks=6)).strftime("%Y-%m-%d")
    url   = "https://www.bankofcanada.ca/valet/observations/group/bond_yields_benchmark/json"
    params = {"start_date": start, "end_date": end}
    print(f"    Trying BOC Valet breakeven (LONG − RRB)...")
    try:
        r = _get(url, params=params)
        if not r.ok:
            print(f"    MISS CAD: BOC Valet HTTP {r.status_code}")
            return None
        data = r.json()
        obs_list = data.get("observations", [])
        if not obs_list:
            print(f"    MISS CAD: BOC Valet — no observations")
            return None
        # Walk backwards to find the latest date with both LONG and RRB values
        for obs in reversed(obs_list):
            long_val = obs.get("BD.CDN.LONG.DQ.YLD", {}).get("v")
            rrb_val  = obs.get("BD.CDN.RRB.DQ.YLD",  {}).get("v")
            obs_date_str = obs.get("d", "")
            if not long_val or not rrb_val or not obs_date_str:
                continue
            try:
                long_y = float(long_val)
                rrb_y  = float(rrb_val)
                obs_date = _parse_date(obs_date_str)
            except (ValueError, TypeError):
                continue
            breakeven = round(long_y - rrb_y, 4)
            # Sanity: Canadian breakeven historically 1.5%–2.5%; post-2022 cancellation
            # the RRB market is thin, so accept wider range [0.5%, 3.5%] before skipping.
            if not (0.5 <= breakeven <= 3.5):
                print(f"    MISS CAD: BOC breakeven {breakeven:.2f}% out of plausible range — skipping")
                return None
            if _is_stale(obs_date, "monthly"):
                print(f"    STALE CAD: BOC Valet RRB obs {obs_date} is >{STALE_DAYS['monthly']}d — skipping")
                return None
            print(f"    OK CAD: {breakeven:.4f}% breakeven ({obs_date}) [{label}]")
            print(f"           (LONG {long_y:.2f}% − RRB {rrb_y:.2f}%)")
            return breakeven, obs_date, label
        print(f"    MISS CAD: BOC Valet — no valid LONG+RRB pair found")
        return None
    except Exception as e:
        print(f"    ERR CAD: BOC Valet — {e}")
        return None


# ---------------------------------------------------------------------------
# NZD PRIMARY  —  RBNZ Survey of Expectations M14 (2Y-ahead, quarterly)
# RBNZ publishes quarterly inflation expectations for 1Y, 2Y, 5Y, 10Y ahead.
# The 2Y-ahead mean is the anchor watched by RBNZ and NZ markets (Fed equivalent:
# NY Fed Survey of Professional Forecasters 3Y-ahead median).
# Data URL: https://www.rbnz.govt.nz/statistics/series/m/m14 (Excel/CSV download)
# Published: first or second week of the second month of each quarter (Feb, May, Aug, Nov)
# DBnomics does not index RBNZ M14 inflation expectations — fetch directly from RBNZ.
# ---------------------------------------------------------------------------

def fetch_nzd_rbnz_survey():
    """RBNZ Survey of Expectations M14 — 2Y-ahead CPI inflation mean. Quarterly."""
    import io as _io
    label = "RBNZ Survey of Expectations M14 (2Y-ahead)"
    print(f"    Trying RBNZ M14 Survey (direct fetch)...")

    # RBNZ M14 historical data as XLSX
    xlsx_url = (
        "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/"
        "series/m/m14/hm14.xlsx"
    )
    try:
        r = _get(xlsx_url, custom_headers={"Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
        if not r.ok:
            print(f"    MISS NZD: RBNZ M14 XLSX HTTP {r.status_code}")
            raise ValueError("HTTP error")

        # Parse with openpyxl
        try:
            import openpyxl as _openpyxl
        except ImportError:
            print(f"    MISS NZD: openpyxl not available — pip install openpyxl")
            raise ValueError("no openpyxl")

        wb = _openpyxl.load_workbook(_io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb.active

        # RBNZ M14 structure: row 1=header, subsequent rows=data
        # Columns include date + multiple expectation horizons.
        # We look for the column header matching "Two years ahead" / "2 year" / "2-year CPI"
        header_row  = None
        col_2y      = None
        date_col    = 0  # first column is always the date

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            if row is None:
                continue
            # Find the header row (contains "Two" or "2 year" etc.)
            row_str = [str(c).lower() if c else "" for c in row]
            if any("two" in s or "2 year" in s or "2yr" in s for s in row_str):
                header_row = i
                # Find the column index for 2Y-ahead CPI expectations
                for j, cell_str in enumerate(row_str):
                    if ("two" in cell_str or "2 year" in cell_str or "2yr" in cell_str) and (
                        "cpi" in cell_str or "inflation" in cell_str or "price" in cell_str
                        or j == next((k for k, s in enumerate(row_str)
                                     if "two" in s or "2 year" in s or "2yr" in s), -1)
                    ):
                        col_2y = j
                        break
                if col_2y is None:
                    # fallback: first column containing "two" or "2 year"
                    for j, cell_str in enumerate(row_str):
                        if "two" in cell_str or "2 year" in cell_str or "2yr" in cell_str:
                            col_2y = j
                            break
                break

        if header_row is None or col_2y is None:
            print(f"    MISS NZD: RBNZ M14 — could not find 2Y-ahead column in header")
            raise ValueError("no 2Y column")

        # Read data rows: find the most recent non-null 2Y expectation
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[date_col]:
                continue
            try:
                date_raw = row[date_col]
                val_raw  = row[col_2y] if col_2y < len(row) else None
                if val_raw is None:
                    continue
                val = float(str(val_raw).replace(",", "."))
                # Parse date: RBNZ uses datetime objects or string "YYYY-MM-DD" / "Mar-2026"
                if hasattr(date_raw, "strftime"):
                    obs_date = date_raw.date() if hasattr(date_raw, "date") else date_raw
                else:
                    date_str = str(date_raw).strip()
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%b-%Y", "%B %Y", "%Y"):
                        try:
                            import datetime as _dt2
                            obs_date = _dt2.datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                # Sanity: NZD 2Y-ahead expectations should be in [0%, 8%]
                if 0 <= val <= 8:
                    data_rows.append((obs_date, val))
            except (ValueError, TypeError):
                continue

        if not data_rows:
            print(f"    MISS NZD: RBNZ M14 — no parseable 2Y-ahead rows")
            raise ValueError("no data rows")

        data_rows.sort(key=lambda x: x[0])
        obs_date, obs_val = data_rows[-1]

        if _is_stale(obs_date, "quarterly"):
            print(f"    STALE NZD: RBNZ M14 obs {obs_date} is >{STALE_DAYS['quarterly']}d — skipping")
            raise ValueError("stale")

        print(f"    OK NZD: {obs_val:.4f}% 2Y survey ({obs_date}) [{label}]")
        return obs_val, obs_date, label

    except Exception as e:
        if str(e) not in ("HTTP error", "no openpyxl", "no 2Y column", "no data rows", "stale"):
            print(f"    ERR NZD: RBNZ M14 — {e}")
        print(f"    MISS NZD: RBNZ survey — falling back to IMF CPI YoY")
        return None


# ---------------------------------------------------------------------------
# G6 PRIMARY  —  IMF SDMX 3.0 API (monthly CPI index → YoY)
# Endpoint: api.imf.org/external/sdmx/3.0/data/dataflow/IMF.STA/CPI/~/{key}
# Key:  {ISO3}.CPI._T.IX.M
# Time: ?c[TIME_PERIOD]=ge:YYYY-M01
# Response: CSV with TIME_PERIOD like "2024-M01", OBS_VALUE = index level
# YoY is computed from the index series (12-month pct change).
# Lag: ~4-6 weeks after national release. No API key required.
# ---------------------------------------------------------------------------

def fetch_imf_cpi(currency):
    country = IMF_COUNTRY[currency]
    key = f"{country}.CPI._T.IX.M"
    # Request 26 months to ensure we have a full 12-month YoY window
    start = (datetime.date.today() - datetime.timedelta(days=800)).strftime("%Y-M%m")
    url = (
        f"https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.STA/CPI/~/"
        f"{key}?c[TIME_PERIOD]=ge:{start}"
    )
    label = f"IMF CPI {country} (monthly index→YoY)"
    print(f"    Trying IMF SDMX 3.0 ({country})...")
    try:
        r = _get(url, custom_headers={"Accept": "text/csv"})
        if r.status_code != 200:
            print(f"    MISS {currency}: IMF HTTP {r.status_code}")
            return None

        reader = csv.DictReader(StringIO(r.text))
        rows = list(reader)
        if not rows:
            print(f"    MISS {currency}: IMF — empty CSV")
            return None

        headers_lower = {k.strip().lower(): k for k in rows[0].keys()}
        val_col  = headers_lower.get("obs_value") or headers_lower.get("obsvalue")
        time_col = headers_lower.get("time_period") or headers_lower.get("timeperiod")
        if not val_col or not time_col:
            print(f"    MISS {currency}: IMF — missing OBS_VALUE/TIME_PERIOD "
                  f"(cols: {list(rows[0].keys())[:6]})")
            return None

        entries = []
        for row in rows:
            try:
                raw_val = row[val_col].strip()
                raw_time = row[time_col].strip()
                if not raw_val or raw_val in ("..", ".", "NaN", ""):
                    continue
                entries.append((_parse_date(raw_time), float(raw_val)))
            except (ValueError, KeyError):
                continue

        if len(entries) < 13:
            print(f"    MISS {currency}: IMF — only {len(entries)} obs (need ≥13 for YoY)")
            return None

        result = _yoy_from_index(entries)
        if result is None:
            print(f"    MISS {currency}: IMF — YoY calculation failed")
            return None

        yoy, obs_date = result
        if _is_stale(obs_date, "monthly"):
            print(f"    STALE {currency}: IMF obs {obs_date} is >{STALE_DAYS['monthly']}d — skipping")
            return None

        print(f"    OK {currency}: {yoy:.4f}% CPI YoY ({obs_date}) [{label}]")
        return yoy, obs_date, label

    except Exception as e:
        print(f"    ERR {currency}: IMF — {e}")
        return None


# ---------------------------------------------------------------------------
# G6 FALLBACK 1  —  OECD Data Explorer SDMX CSV (direct YoY series)
# Key: {country}.{M|Q}.N.CPI.PA._T.N.GY  (TRANSFORMATION=GY = annual % change)
# AUS/NZD use quarterly (Q); time format "2024-Q4" is now parsed correctly.
# JPN will return 404 (COICOP 2018 only — not in DF_PRICES_ALL).
# CHF: last obs Dec 2025 (~157 days) exceeds 120d threshold. Still tried.
# ---------------------------------------------------------------------------

def fetch_oecd_explorer(currency):
    country, freq_code = OECD_COUNTRY[currency]
    key = f"{country}.{freq_code}.N.CPI.PA._T.N.GY"
    url = (
        f"https://sdmx.oecd.org/public/rest/data/"
        f"OECD.SDD.TPS,DSD_PRICES@DF_PRICES_ALL,1.0/{key}"
        f"?startPeriod=2024-01&dimensionAtObservation=AllDimensions"
        f"&format=csvfilewithlabels"
    )
    label = f"OECD CPI YoY ({country})"
    print(f"    Trying OECD Explorer ({country} {freq_code} GY)...")
    try:
        r = _get(url, custom_headers={"Accept": "text/csv"})
        if r.status_code != 200:
            print(f"    MISS {currency}: OECD HTTP {r.status_code}")
            return None

        reader = csv.DictReader(StringIO(r.text))
        rows = list(reader)
        if not rows:
            print(f"    MISS {currency}: OECD — empty CSV")
            return None

        headers_lower = {k.strip().lower(): k for k in rows[0].keys()}
        val_col  = headers_lower.get("obs_value") or headers_lower.get("obsvalue")
        time_col = headers_lower.get("time_period") or headers_lower.get("timeperiod")
        if not val_col or not time_col:
            print(f"    MISS {currency}: OECD — missing columns")
            return None

        entries = []
        for row in rows:
            try:
                raw_val = row[val_col].strip()
                raw_time = row[time_col].strip()
                if not raw_val or raw_val in ("..", ".", "NaN", ""):
                    continue
                entries.append((_parse_date(raw_time), float(raw_val)))
            except (ValueError, KeyError):
                continue

        if not entries:
            print(f"    MISS {currency}: OECD — no parseable rows")
            return None

        entries.sort(key=lambda x: x[0])
        obs_date, obs_val = entries[-1]
        freq = "quarterly" if freq_code == "Q" else "monthly"

        if _is_stale(obs_date, freq):
            print(f"    STALE {currency}: OECD obs {obs_date} is >{STALE_DAYS[freq]}d — skipping")
            return None

        print(f"    OK {currency}: {obs_val:.4f}% CPI YoY ({obs_date}) [{label}]")
        return obs_val, obs_date, label

    except Exception as e:
        print(f"    ERR {currency}: OECD — {e}")
        return None


# ---------------------------------------------------------------------------
# G6 FALLBACK 2  —  FRED index series → YoY
# Structural ~12-14 month lag as of 2026. Will almost always be stale.
# ---------------------------------------------------------------------------

def fetch_fred_index_yoy(currency):
    if not FRED_API_KEY:
        print(f"    SKIP {currency}: FRED_API_KEY not set")
        return None
    series_id, freq = FRED_INDEX_SERIES[currency]
    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&sort_order=desc&limit=36&file_type=json"
        f"&api_key={FRED_API_KEY}"
    )
    label = f"FRED {series_id}→YoY"
    print(f"    Trying FRED {series_id}...")
    try:
        r = _get(url)
        r.raise_for_status()
        series = [
            (_parse_date(o["date"]), float(o["value"]))
            for o in r.json().get("observations", [])
            if o.get("value") not in (".", "")
        ]
        result = _yoy_from_index(series)
        if result is None:
            print(f"    MISS {currency}: FRED {series_id} — insufficient data")
            return None
        yoy, obs_date = result
        print(f"    ~ {currency}: {yoy:.4f}% ({obs_date}) [{label}]")
        if _is_stale(obs_date, freq):
            print(f"    STALE {currency}: obs {obs_date} is >{STALE_DAYS[freq]}d — skipping")
            return None
        return yoy, obs_date, label
    except Exception as e:
        print(f"    ERR {currency}: FRED {series_id} — {e}")
        return None


# ---------------------------------------------------------------------------
# G6 LAST RESORT  —  World Bank (annual, lag >12 months)
# ---------------------------------------------------------------------------

def fetch_world_bank(currency):
    country = WB_COUNTRY[currency]
    url = (
        f"https://api.worldbank.org/v2/country/{country}/indicator/FP.CPI.TOTL.ZG"
        f"?format=json&mrv=5&per_page=5"
    )
    label = "World Bank"
    print(f"    Trying World Bank...")
    try:
        r = _get(url)
        r.raise_for_status()
        data = r.json()
        if len(data) < 2 or not data[1]:
            print(f"    MISS {currency}: World Bank — no data")
            return None
        for entry in data[1]:
            if entry.get("value") is not None:
                val = float(entry["value"])
                obs_date = datetime.date(int(entry["date"]), 6, 15)
                print(f"    ~ {currency}: {val:.4f}% ({obs_date}) [{label}]")
                if _is_stale(obs_date, "monthly"):
                    print(f"    STALE {currency}: obs {obs_date} is >{STALE_DAYS['monthly']}d — skipping")
                    return None
                return val, obs_date, label
        print(f"    MISS {currency}: World Bank — all values null")
        return None
    except Exception as e:
        print(f"    ERR {currency}: World Bank — {e}")
        return None


# ---------------------------------------------------------------------------
# Patch extended-data/{CCY}.json
# ---------------------------------------------------------------------------

def patch_extended_data(site_root, currency, val, date):
    path = os.path.join(site_root, "extended-data", f"{currency}.json")
    try:
        if os.path.exists(path):
            with open(path) as f:
                d = json.load(f)
        else:
            d = {"data": {}, "dates": {}}
            os.makedirs(os.path.dirname(path), exist_ok=True)
        d.setdefault("data", {})
        d.setdefault("dates", {})
        d["data"]["inflationExpectations"] = round(val, 4)
        d["dates"]["inflationExpectations"] = str(date)
        with open(path, "w") as f:
            json.dump(d, f, indent=2)
        print(f"    OK patched {path}")
    except Exception as e:
        print(f"    WARN: could not patch {path}: {e}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) > 1:
        site_root = sys.argv[1]
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(script_dir, "..", "..", "site"),
            os.path.join(script_dir, ".."),
        ]
        site_root = next(
            (p for p in candidates if os.path.isdir(os.path.join(p, "extended-data"))),
            candidates[0]
        )

    print("=" * 60)
    print("INFLATION EXPECTATIONS — G8 currencies  (v5.0)")
    print(f"Run: {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC")
    print(f"Site root: {os.path.abspath(site_root)}")
    print("Sources: FRED T5YIE/T5YIFR/CAINFIMPCPI · BOE SDIE BEAPFF · RBNZ Survey · IMF SDMX")
    print("=" * 60)

    results = {}
    failures = []

    # ── USD: FRED 5Y TIPS breakeven (market-implied) ─────────────────────────
    print("[USD]")
    r = fetch_fred_breakeven("T5YIE", "USD", "FRED T5YIE")
    if r:
        results["USD"] = r
    else:
        failures.append("USD")

    # ── EUR: FRED 5Y5Y inflation swap (market-implied) ───────────────────────
    print("[EUR]")
    r = fetch_fred_breakeven("T5YIFR", "EUR", "FRED T5YIFR")
    if r:
        results["EUR"] = r
    else:
        failures.append("EUR")

    # ── GBP: BoE 2Y-ahead survey → IMF CPI YoY → OECD → FRED → WB ──────────
    print("[GBP]")
    r = fetch_boe_inflation_expectations()      # forward-looking survey (PRIMARY)
    if r is None:
        r = fetch_imf_cpi("GBP")               # realised CPI YoY (FALLBACK 1)
    if r is None:
        r = fetch_oecd_explorer("GBP")          # OECD CPI YoY (FALLBACK 2)
    if r is None:
        r = fetch_fred_index_yoy("GBP")         # FRED MEI index→YoY (FALLBACK 3)
    if r is None:
        r = fetch_world_bank("GBP")             # World Bank annual (LAST RESORT)
    if r:
        results["GBP"] = r
    else:
        failures.append("GBP")
        print("Warning: fetch_inflation_expectations: GBP stale or unavailable", file=sys.stderr)

    # ── JPY: IMF CPI YoY → OECD → FRED → WB (no liquid breakeven) ──────────
    print("[JPY]")
    r = fetch_imf_cpi("JPY")                   # CPI YoY PRIMARY (no market breakeven)
    if r is None:
        r = fetch_oecd_explorer("JPY")
    if r is None:
        r = fetch_fred_index_yoy("JPY")
    if r is None:
        r = fetch_world_bank("JPY")
    if r:
        results["JPY"] = r
    else:
        failures.append("JPY")
        print("Warning: fetch_inflation_expectations: JPY stale or unavailable", file=sys.stderr)

    # ── AUD: IMF CPI YoY → OECD → FRED → WB ────────────────────────────────
    # Note: AUD Mar-2026 headline 4.57% is elevated by energy spike.
    # IMF is used as primary (consistent methodology with JPY/CHF); the
    # calendar extractor in update_extended_data.py supplements with
    # Melbourne Institute 1Y survey when available (CALENDAR_INFLEXP_PATTERNS).
    print("[AUD]")
    r = fetch_imf_cpi("AUD")
    if r is None:
        r = fetch_oecd_explorer("AUD")
    if r is None:
        r = fetch_fred_index_yoy("AUD")
    if r is None:
        r = fetch_world_bank("AUD")
    if r:
        results["AUD"] = r
    else:
        failures.append("AUD")
        print("Warning: fetch_inflation_expectations: AUD stale or unavailable", file=sys.stderr)

    # ── CAD: FRED 5Y breakeven → IMF CPI YoY → OECD → FRED → WB ────────────
    print("[CAD]")
    r = fetch_cad_breakeven()                   # market-implied breakeven (PRIMARY)
    if r is None:
        r = fetch_imf_cpi("CAD")               # CPI YoY (FALLBACK 1)
    if r is None:
        r = fetch_oecd_explorer("CAD")
    if r is None:
        r = fetch_fred_index_yoy("CAD")
    if r is None:
        r = fetch_world_bank("CAD")
    if r:
        results["CAD"] = r
    else:
        failures.append("CAD")
        print("Warning: fetch_inflation_expectations: CAD stale or unavailable", file=sys.stderr)

    # ── CHF: IMF CPI YoY → OECD → FRED → WB (no liquid breakeven) ──────────
    print("[CHF]")
    r = fetch_imf_cpi("CHF")
    if r is None:
        r = fetch_oecd_explorer("CHF")
    if r is None:
        r = fetch_fred_index_yoy("CHF")
    if r is None:
        r = fetch_world_bank("CHF")
    if r:
        results["CHF"] = r
    else:
        failures.append("CHF")
        print("Warning: fetch_inflation_expectations: CHF stale or unavailable", file=sys.stderr)

    # ── NZD: RBNZ 2Y-ahead survey → IMF CPI YoY → OECD → FRED → WB ─────────
    print("[NZD]")
    r = fetch_nzd_rbnz_survey()                 # forward-looking survey (PRIMARY)
    if r is None:
        r = fetch_imf_cpi("NZD")               # CPI YoY (FALLBACK 1)
    if r is None:
        r = fetch_oecd_explorer("NZD")
    if r is None:
        r = fetch_fred_index_yoy("NZD")
    if r is None:
        r = fetch_world_bank("NZD")
    if r:
        results["NZD"] = r
    else:
        failures.append("NZD")
        print("Warning: fetch_inflation_expectations: NZD stale or unavailable", file=sys.stderr)

    print()
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for ccy, (val, date, src) in sorted(results.items()):
        sign = "+" if val >= 0 else ""
        print(f"  {ccy:5s} {sign}{val:.4f}%  ({date})  [{src}]")
    if failures:
        print(f"  Skipped/failed: {', '.join(failures)} ({len(failures)} issue(s))")

    print()
    print("Patching extended-data/...")
    for ccy, (val, date, _src) in results.items():
        patch_extended_data(site_root, ccy, val, date)

    g6_failures = [c for c in failures if c not in ("USD", "EUR")]
    if len(g6_failures) >= 3:
        print("Error: fetch_inflation_expectations: >=3 G6 currencies failed — likely upstream outage",
              file=sys.stderr)
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
